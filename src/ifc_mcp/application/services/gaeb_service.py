"""GAEB Service (German Construction Tender Standard).

Service for generating GAEB X84 XML and Excel exports.
Migrated from cad_hub with proper async architecture.
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from io import BytesIO
from typing import TYPE_CHECKING
from uuid import UUID

if TYPE_CHECKING:
    from ifc_mcp.infrastructure.repositories.unit_of_work import UnitOfWork


class GAEBPhase(str, Enum):
    """GAEB data exchange phases."""

    X81 = "81"  # Request
    X83 = "83"  # Offer (with prices)
    X84 = "84"  # Alternative offer
    X85 = "85"  # Contract award


class MengenEinheit(str, Enum):
    """Quantity units according to GAEB."""

    STK = "Stk"  # Piece
    M = "m"  # Meter
    M2 = "m2"  # Square meter
    M3 = "m3"  # Cubic meter
    KG = "kg"  # Kilogram
    T = "t"  # Ton
    L = "l"  # Liter
    H = "h"  # Hour
    PAU = "psch"  # Lump sum
    TAG = "Tag"  # Day


@dataclass
class Position:
    """Bill of quantities position."""

    oz: str  # Position number (e.g., "01.02.0010")
    kurztext: str  # Short description
    langtext: str = ""  # Long description
    menge: Decimal = Decimal("0")  # Quantity
    einheit: MengenEinheit = MengenEinheit.STK  # Unit
    einheitspreis: Decimal = Decimal("0")  # Unit price
    gesamtpreis: Decimal = Decimal("0")  # Total price
    stlb_code: str = ""  # STLB-Bau reference

    def __post_init__(self) -> None:
        """Calculate total price if not provided."""
        if self.gesamtpreis == 0 and self.menge > 0 and self.einheitspreis > 0:
            self.gesamtpreis = self.menge * self.einheitspreis

    def to_dict(self) -> dict:
        """Convert to dict."""
        return {
            "oz": self.oz,
            "kurztext": self.kurztext,
            "menge": float(self.menge),
            "einheit": self.einheit.value,
            "einheitspreis": float(self.einheitspreis),
            "gesamtpreis": float(self.gesamtpreis),
        }


@dataclass
class LosGruppe:
    """Lot/Title/Group in bill of quantities."""

    oz: str  # Group number
    bezeichnung: str  # Description
    positionen: list[Position] = field(default_factory=list)
    untergruppen: list[LosGruppe] = field(default_factory=list)

    @property
    def summe(self) -> Decimal:
        """Total sum of all positions and subgroups."""
        pos_summe = sum((p.gesamtpreis for p in self.positionen), Decimal("0"))
        ug_summe = sum((ug.summe for ug in self.untergruppen), Decimal("0"))
        return pos_summe + ug_summe

    @property
    def anzahl_positionen(self) -> int:
        """Total number of positions including subgroups."""
        return len(self.positionen) + sum(
            ug.anzahl_positionen for ug in self.untergruppen
        )


@dataclass
class Leistungsverzeichnis:
    """Complete bill of quantities."""

    projekt_name: str
    projekt_nummer: str = ""
    lv_nummer: str = ""
    auftraggeber: str = ""
    auftragnehmer: str = ""
    lose: list[LosGruppe] = field(default_factory=list)
    waehrung: str = "EUR"
    datum: date = field(default_factory=date.today)
    phase: GAEBPhase = GAEBPhase.X83

    @property
    def netto_summe(self) -> Decimal:
        """Net total."""
        return sum((los.summe for los in self.lose), Decimal("0"))

    @property
    def mwst(self) -> Decimal:
        """VAT (19%)."""
        return self.netto_summe * Decimal("0.19")

    @property
    def brutto_summe(self) -> Decimal:
        """Gross total."""
        return self.netto_summe + self.mwst

    @property
    def anzahl_positionen(self) -> int:
        """Total number of positions."""
        return sum(los.anzahl_positionen for los in self.lose)


class GAEBService:
    """Service for GAEB X84 generation.

    Generates German construction tender documents (Leistungsverzeichnis).
    """

    GAEB_NAMESPACE = "http://www.gaeb.de/GAEB_DA_XML/200407"

    def __init__(self, uow: UnitOfWork) -> None:
        """Initialize service.

        Args:
            uow: Unit of Work instance
        """
        self._uow = uow

    async def generate_from_project(
        self,
        project_id: UUID,
        projekt_nummer: str = "",
    ) -> Leistungsverzeichnis:
        """Generate bill of quantities from project.

        Args:
            project_id: Project UUID
            projekt_nummer: Project number (optional)

        Returns:
            Complete Leistungsverzeichnis
        """
        # Get project
        project = await self._uow.projects.get(project_id)
        if not project:
            raise ValueError(f"Project {project_id} not found")

        # Get spaces
        spaces = await self._uow.spaces.find_by_project(project_id)

        # Create LV
        lv = Leistungsverzeichnis(
            projekt_name=project.name,
            projekt_nummer=projekt_nummer or str(project_id)[:8],
        )

        # Los 1: Floor coverings
        if spaces:
            boden_positionen = []
            for idx, space in enumerate(spaces, 1):
                pos = Position(
                    oz=f"01.01.{idx:04d}",
                    kurztext=f"Bodenbelag {space.number or ''} {space.name or ''}".strip(),
                    menge=space.net_floor_area_m2 or Decimal("0"),
                    einheit=MengenEinheit.M2,
                )
                boden_positionen.append(pos)

            if boden_positionen:
                lv.lose.append(
                    LosGruppe(
                        oz="01",
                        bezeichnung="Bodenbeläge",
                        positionen=boden_positionen,
                    )
                )

        return lv

    def generate_xml(self, lv: Leistungsverzeichnis) -> BytesIO:
        """Generate GAEB X84 XML.

        Args:
            lv: Leistungsverzeichnis

        Returns:
            BytesIO with XML content
        """
        root = self._create_root(lv)
        self._indent(root)

        output = BytesIO()
        tree = ET.ElementTree(root)
        tree.write(output, encoding="utf-8", xml_declaration=True)
        output.seek(0)
        return output

    def generate_excel(self, lv: Leistungsverzeichnis) -> BytesIO:
        """Generate Excel bill of quantities.

        Args:
            lv: Leistungsverzeichnis

        Returns:
            BytesIO with Excel content
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Leistungsverzeichnis"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="003366", end_color="003366", fill_type="solid"
        )
        sum_font = Font(bold=True)
        sum_fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )

        # Project info
        ws["A1"] = f"Leistungsverzeichnis: {lv.projekt_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Projekt-Nr.: {lv.projekt_nummer}" if lv.projekt_nummer else ""
        ws["A3"] = f"Datum: {lv.datum.strftime('%d.%m.%Y')}"

        # Header (row 5)
        headers = ["OZ", "Kurztext", "Menge", "Einheit", "EP [\u20ac]", "GP [\u20ac]"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12

        # Data
        row = 6
        for los in lv.lose:
            row = self._write_gruppe_excel(ws, los, row, sum_font, sum_fill)

        # Totals
        row += 1
        ws.cell(row=row, column=2, value="NETTO SUMME:").font = sum_font
        ws.cell(row=row, column=6, value=float(lv.netto_summe)).font = sum_font
        ws.cell(row=row, column=6).number_format = "#,##0.00 \u20ac"

        row += 1
        ws.cell(row=row, column=2, value="MwSt 19%:")
        ws.cell(row=row, column=6, value=float(lv.mwst))
        ws.cell(row=row, column=6).number_format = "#,##0.00 \u20ac"

        row += 1
        ws.cell(row=row, column=2, value="BRUTTO SUMME:").font = sum_font
        cell = ws.cell(row=row, column=6, value=float(lv.brutto_summe))
        cell.font = sum_font
        cell.fill = sum_fill
        cell.number_format = "#,##0.00 \u20ac"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_gruppe_excel(
        self, ws, gruppe: LosGruppe, row: int, sum_font, sum_fill
    ) -> int:
        """Write group to Excel."""
        from openpyxl.styles import Font

        # Group row
        ws.cell(row=row, column=1, value=gruppe.oz).font = Font(bold=True)
        ws.cell(row=row, column=2, value=gruppe.bezeichnung).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(gruppe.summe)).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = "#,##0.00 \u20ac"
        row += 1

        # Positions
        for pos in gruppe.positionen:
            ws.cell(row=row, column=1, value=pos.oz)
            ws.cell(row=row, column=2, value=pos.kurztext)
            ws.cell(row=row, column=3, value=float(pos.menge))
            ws.cell(row=row, column=4, value=pos.einheit.value)
            ws.cell(row=row, column=5, value=float(pos.einheitspreis))
            ws.cell(row=row, column=5).number_format = "#,##0.00"
            ws.cell(row=row, column=6, value=float(pos.gesamtpreis))
            ws.cell(row=row, column=6).number_format = "#,##0.00"
            row += 1

        # Subgroups
        for ug in gruppe.untergruppen:
            row = self._write_gruppe_excel(ws, ug, row, sum_font, sum_fill)

        return row

    def _create_root(self, lv: Leistungsverzeichnis) -> ET.Element:
        """Create GAEB XML root."""
        root = ET.Element("GAEB")
        root.set("xmlns", self.GAEB_NAMESPACE)

        # GAEBInfo
        gaeb_info = ET.SubElement(root, "GAEBInfo")
        ET.SubElement(gaeb_info, "Version").text = "GAEB XML 3.2"
        ET.SubElement(gaeb_info, "VersNo").text = "32"
        ET.SubElement(gaeb_info, "Date").text = datetime.now().isoformat()
        ET.SubElement(gaeb_info, "ProgSystem").text = "IFC MCP"

        # PrjInfo
        prj_info = ET.SubElement(root, "PrjInfo")
        ET.SubElement(prj_info, "NamePrj").text = lv.projekt_name
        if lv.projekt_nummer:
            ET.SubElement(prj_info, "LblPrj").text = lv.projekt_nummer
        ET.SubElement(prj_info, "Cur").text = lv.waehrung

        # Award
        award = ET.SubElement(root, "Award")
        boq = ET.SubElement(award, "BoQ")
        ET.SubElement(boq, "BoQInfo")
        boq_body = ET.SubElement(boq, "BoQBody")

        for los in lv.lose:
            self._add_gruppe(boq_body, los)

        return root

    def _add_gruppe(self, parent: ET.Element, gruppe: LosGruppe) -> None:
        """Add group to XML."""
        boq_ctgy = ET.SubElement(parent, "BoQCtgy")
        ET.SubElement(boq_ctgy, "LblTx").text = gruppe.oz
        ET.SubElement(boq_ctgy, "Headline").text = gruppe.bezeichnung

        boq_body = ET.SubElement(boq_ctgy, "BoQBody")

        for pos in gruppe.positionen:
            self._add_position(boq_body, pos)

        for ug in gruppe.untergruppen:
            self._add_gruppe(boq_body, ug)

    def _add_position(self, parent: ET.Element, pos: Position) -> None:
        """Add position to XML."""
        item = ET.SubElement(parent, "Itemlist")
        item_elem = ET.SubElement(item, "Item")

        ET.SubElement(item_elem, "Qty").text = str(pos.menge)
        ET.SubElement(item_elem, "QU").text = pos.einheit.value

        description = ET.SubElement(item_elem, "Description")
        ET.SubElement(description, "OutlineText").text = pos.kurztext
        if pos.langtext:
            ET.SubElement(description, "DetailTxt").text = pos.langtext

    def _indent(self, elem: ET.Element, level: int = 0) -> None:
        """Add XML indentation."""
        i = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for child in elem:
                self._indent(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i
